import pytest
import pandas as pd
from pathlib import Path
from app.services.export_service import ExportService


class TestExportService:
    
    @pytest.fixture(autouse=True)
    def setup(self, test_data_dir):
        self.service = ExportService()
        self.service.output_dir = test_data_dir
    
    def test_init(self):
        assert self.service.output_dir.exists()
    
    def test_export_to_excel_default_filename(self, sample_dataframe):
        result_path = self.service.export_to_excel(sample_dataframe)
        
        assert Path(result_path).exists()
        assert result_path.endswith('.xlsx')
        
        df_read = pd.read_excel(result_path)
        assert len(df_read) == len(sample_dataframe)
    
    def test_export_to_excel_custom_filename(self, sample_dataframe):
        result_path = self.service.export_to_excel(sample_dataframe, filename='custom.xlsx')
        
        assert 'custom.xlsx' in result_path
        assert Path(result_path).exists()
    
    def test_export_to_excel_auto_extension(self, sample_dataframe):
        result_path = self.service.export_to_excel(sample_dataframe, filename='noextension')
        
        assert result_path.endswith('.xlsx')
        assert Path(result_path).exists()
    
    def test_export_to_excel_custom_sheet_name(self, sample_dataframe):
        result_path = self.service.export_to_excel(
            sample_dataframe, 
            filename='sheet_test.xlsx',
            sheet_name='CustomSheet'
        )
        
        xl_file = pd.ExcelFile(result_path)
        assert 'CustomSheet' in xl_file.sheet_names
    
    def test_export_to_excel_with_index(self, sample_dataframe):
        result_path = self.service.export_to_excel(
            sample_dataframe,
            filename='with_index.xlsx',
            include_index=True
        )
        
        df_read = pd.read_excel(result_path)
        assert df_read.columns[0] == 'Unnamed: 0'
    
    def test_export_multiple_sheets(self, sample_dataframe):
        df2 = sample_dataframe.copy()
        df2['extra_col'] = 999
        
        dataframes = {
            'Sheet1': sample_dataframe,
            'Sheet2': df2
        }
        
        result_path = self.service.export_multiple_sheets(dataframes, 'multi_sheet.xlsx')
        
        assert Path(result_path).exists()
        
        xl_file = pd.ExcelFile(result_path)
        assert 'Sheet1' in xl_file.sheet_names
        assert 'Sheet2' in xl_file.sheet_names
    
    def test_export_multiple_sheets_auto_extension(self, sample_dataframe):
        dataframes = {'Sheet1': sample_dataframe}
        
        result_path = self.service.export_multiple_sheets(dataframes, 'noext')
        
        assert result_path.endswith('.xlsx')
    
    def test_export_with_formatting(self, sample_dataframe):
        result_path = self.service.export_with_formatting(
            sample_dataframe,
            filename='formatted.xlsx'
        )
        
        assert Path(result_path).exists()
        
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        ws = wb.active
        
        header_cell = ws['A1']
        assert header_cell.font.bold is True
    
    def test_export_with_formatting_custom_sheet(self, sample_dataframe):
        result_path = self.service.export_with_formatting(
            sample_dataframe,
            filename='formatted_custom.xlsx',
            sheet_name='MySheet'
        )
        
        xl_file = pd.ExcelFile(result_path)
        assert 'MySheet' in xl_file.sheet_names
    
    def test_export_to_csv_default_filename(self, sample_dataframe):
        result_path = self.service.export_to_csv(sample_dataframe)
        
        assert Path(result_path).exists()
        assert result_path.endswith('.csv')
        
        df_read = pd.read_csv(result_path)
        assert len(df_read) == len(sample_dataframe)
    
    def test_export_to_csv_custom_filename(self, sample_dataframe):
        result_path = self.service.export_to_csv(sample_dataframe, filename='custom.csv')
        
        assert 'custom.csv' in result_path
        assert Path(result_path).exists()
    
    def test_export_to_csv_auto_extension(self, sample_dataframe):
        result_path = self.service.export_to_csv(sample_dataframe, filename='noext')
        
        assert result_path.endswith('.csv')
    
    def test_export_to_csv_with_index(self, sample_dataframe):
        result_path = self.service.export_to_csv(
            sample_dataframe,
            filename='with_index.csv',
            include_index=True
        )
        
        df_read = pd.read_csv(result_path)
        assert df_read.columns[0] == 'Unnamed: 0'
    
    def test_get_export_summary(self, sample_dataframe):
        summary = self.service.get_export_summary(sample_dataframe)
        
        assert 'rows' in summary
        assert 'columns' in summary
        assert 'column_names' in summary
        assert 'memory_usage_mb' in summary
        assert 'null_counts' in summary
        assert 'dtypes' in summary
        
        assert summary['rows'] == len(sample_dataframe)
        assert summary['columns'] == len(sample_dataframe.columns)
        assert isinstance(summary['column_names'], list)
        assert isinstance(summary['memory_usage_mb'], float)
    
    def test_get_export_summary_with_nulls(self):
        df = pd.DataFrame({
            'a': [1, 2, None, 4],
            'b': [5, None, None, 8]
        })
        
        summary = self.service.get_export_summary(df)
        
        assert summary['null_counts']['a'] == 1
        assert summary['null_counts']['b'] == 2
    
    def test_sanitize_sheet_name_special_chars(self):
        name = "Sheet[1]:Test*Name?"
        sanitized = self.service._sanitize_sheet_name(name)
        
        assert '[' not in sanitized
        assert ']' not in sanitized
        assert ':' not in sanitized
        assert '*' not in sanitized
        assert '?' not in sanitized
    
    def test_sanitize_sheet_name_long_name(self):
        name = "A" * 50
        sanitized = self.service._sanitize_sheet_name(name)
        
        assert len(sanitized) <= 31
    
    def test_sanitize_sheet_name_slash(self):
        name = "Sheet/Name\\Test"
        sanitized = self.service._sanitize_sheet_name(name)
        
        assert '/' not in sanitized
        assert '\\' not in sanitized
    
    def test_list_exports_empty(self):
        exports = self.service.list_exports()
        
        assert isinstance(exports, list)
    
    def test_list_exports_with_files(self, sample_dataframe):
        self.service.export_to_excel(sample_dataframe, filename='test1.xlsx')
        self.service.export_to_excel(sample_dataframe, filename='test2.xlsx')
        
        exports = self.service.list_exports()
        
        assert len(exports) >= 2
        assert any('test1.xlsx' in e['filename'] for e in exports)
        assert any('test2.xlsx' in e['filename'] for e in exports)
    
    def test_list_exports_metadata(self, sample_dataframe):
        self.service.export_to_excel(sample_dataframe, filename='metadata_test.xlsx')
        
        exports = self.service.list_exports()
        
        export = next(e for e in exports if 'metadata_test.xlsx' in e['filename'])
        assert 'filepath' in export
        assert 'size_mb' in export
        assert 'created' in export
        assert isinstance(export['size_mb'], float)
    
    def test_export_empty_dataframe(self):
        df = pd.DataFrame()
        
        result_path = self.service.export_to_excel(df, filename='empty.xlsx')
        
        assert Path(result_path).exists()
    
    def test_export_large_dataframe(self, sample_dataframe):
        large_df = pd.concat([sample_dataframe] * 10, ignore_index=True)
        
        result_path = self.service.export_to_excel(large_df, filename='large.xlsx')
        
        assert Path(result_path).exists()
        
        df_read = pd.read_excel(result_path)
        assert len(df_read) == len(large_df)